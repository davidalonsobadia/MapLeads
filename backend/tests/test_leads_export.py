"""Tests for exporting a project's leads to CSV and XLSX.

The ``client`` fixture is authenticated as ``test_user`` (see ``conftest.py``).
Exports honor the same ``status``/``q`` filters as the leads list and enforce
project ownership.
"""

import csv
import io

from openpyxl import load_workbook

from app.domains.auth.models import User
from app.domains.leads.models import Lead
from app.domains.projects.models import Project

PROJECTS = "/api/v1/projects"
EXPECTED_HEADERS = ["name", "address", "phone", "website", "category", "status", "date saved"]


def _create_project(client, name="Prospects") -> int:
    return client.post(PROJECTS, json={"name": name}).json()["id"]


def _item(place_id, name, **extra):
    return {"place_id": place_id, "name": name, **extra}


def _seed(client):
    """Seed three leads, two marked contacted, and return the leads base URL."""
    project_id = _create_project(client)
    base = f"{PROJECTS}/{project_id}/leads"
    client.post(
        base,
        json={
            "items": [
                _item("p1", "Downtown Diner", address="1 Main St", phone="111"),
                _item("p2", "Uptown Deli", website="https://deli.example"),
                _item("p3", "Riverside Diner", category="Restaurant"),
            ]
        },
    )
    leads = {lead["place_id"]: lead["id"] for lead in client.get(base).json()}
    client.patch(f"/api/v1/leads/{leads['p1']}", json={"status": "contacted"})
    client.patch(f"/api/v1/leads/{leads['p3']}", json={"status": "contacted"})
    return base


def _parse_csv(response):
    reader = csv.reader(io.StringIO(response.content.decode("utf-8-sig")))
    return list(reader)


def test_export_csv_has_headers_and_all_rows(client):
    base = _seed(client)
    response = client.get(f"{base}/export?format=csv")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/csv")
    disposition = response.headers["content-disposition"]
    assert disposition.startswith("attachment;")
    assert disposition.endswith('.csv"')

    rows = _parse_csv(response)
    assert rows[0] == EXPECTED_HEADERS
    # header + 3 data rows
    assert len(rows) == 4
    names = {row[0] for row in rows[1:]}
    assert names == {"Downtown Diner", "Uptown Deli", "Riverside Diner"}


def test_export_csv_defaults_to_csv_without_format(client):
    base = _seed(client)
    response = client.get(f"{base}/export")
    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/csv")


def test_export_csv_honors_status_and_q_filters(client):
    base = _seed(client)

    # status filter -> only the two contacted leads
    contacted = _parse_csv(client.get(f"{base}/export?status=contacted"))
    assert len(contacted) == 3  # header + 2
    assert {row[0] for row in contacted[1:]} == {"Downtown Diner", "Riverside Diner"}

    # name search -> both diners
    diners = _parse_csv(client.get(f"{base}/export?q=diner"))
    assert {row[0] for row in diners[1:]} == {"Downtown Diner", "Riverside Diner"}

    # combined status + name search -> a single row
    combined = client.get(f"{base}/export?status=contacted&q=downtown")
    rows = _parse_csv(combined)
    assert len(rows) == 2
    assert rows[1][0] == "Downtown Diner"


def test_export_xlsx_parses_and_matches_filter(client):
    base = _seed(client)
    response = client.get(f"{base}/export?format=xlsx&status=contacted")

    assert response.status_code == 200
    assert response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert response.headers["content-disposition"].endswith('.xlsx"')

    workbook = load_workbook(io.BytesIO(response.content))
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    assert list(rows[0]) == EXPECTED_HEADERS
    # header + 2 contacted rows
    assert len(rows) == 3
    assert {row[0] for row in rows[1:]} == {"Downtown Diner", "Riverside Diner"}


def test_export_filename_reflects_filters(client):
    base = _seed(client)
    disposition = client.get(
        f"{base}/export?format=xlsx&status=contacted&q=diner"
    ).headers["content-disposition"]
    assert "status-contacted" in disposition
    assert "q-diner" in disposition
    assert disposition.endswith('.xlsx"')


def test_export_rejects_invalid_format(client):
    base = _seed(client)
    # Constrained by the Literal query type, so FastAPI rejects with 422.
    assert client.get(f"{base}/export?format=pdf").status_code == 422


def test_export_ownership_is_enforced(client, db_session):
    """Exporting another user's project returns 404."""
    other_user = User(
        name="Other User",
        email="other@example.com",
        hashed_password="not-a-real-hash",
        is_verified=True,
    )
    db_session.add(other_user)
    db_session.commit()
    db_session.refresh(other_user)

    other_project = Project(user_id=other_user.id, name="Other's Project")
    db_session.add(other_project)
    db_session.commit()
    db_session.refresh(other_project)

    other_lead = Lead(
        project_id=other_project.id,
        user_id=other_user.id,
        place_id="secret",
        name="Confidential",
    )
    db_session.add(other_lead)
    db_session.commit()

    response = client.get(
        f"{PROJECTS}/{other_project.id}/leads/export?format=csv"
    )
    assert response.status_code == 404
